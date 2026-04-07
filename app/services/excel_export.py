"""Excel export for sales, purchases, and inventory."""
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io


HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill('solid', fgColor='1a1a2e')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CURRENCY_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.border = THIN_BORDER
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def export_sales(sales):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales'

    ws.append(['Invoice #', 'Date', 'Customer', 'Type', 'Product', 'Variant',
               'Qty', 'Unit Price', 'GST%', 'GST Amt', 'Total', 'Transport',
               'Discount', 'Grand Total', 'Package', 'Status'])
    style_header(ws)

    for sale in sales:
        first_item = True
        for item in sale.items:
            ws.append([
                sale.invoice_number if first_item else '',
                str(sale.sale_date) if first_item else '',
                sale.customer.name if first_item else '',
                sale.customer.customer_type if first_item else '',
                item.variant.product.name,
                item.variant.display_name,
                item.quantity,
                item.unit_price,
                item.gst_percent,
                item.gst_amount,
                item.total_amount,
                sale.transport_charge if first_item else '',
                sale.discount_amount if first_item else '',
                sale.grand_total if first_item else '',
                sale.package_type or '' if first_item else '',
                sale.status if first_item else '',
            ])
            row = ws.max_row
            for col in [8, 10, 11, 12, 13, 14]:
                ws.cell(row=row, column=col).number_format = CURRENCY_FMT
            first_item = False

    # Summary row
    ws.append([])
    summary_row = ws.max_row + 1
    ws.append(['', '', '', '', '', 'TOTAL', '', '', '', '', '', '', '',
               sum(s.grand_total for s in sales)])
    ws.cell(row=summary_row, column=14).number_format = CURRENCY_FMT
    ws.cell(row=summary_row, column=14).font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=6).font = Font(bold=True, size=11)

    auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='Sportspedal_Sales.xlsx', as_attachment=True)


def export_purchases(purchases):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Purchases'

    ws.append(['Order #', 'Date', 'Supplier', 'Product', 'Variant', 'SKU',
               'Qty Dispatched', 'Qty Received', 'Unit Price', 'GST%',
               'GST Amt', 'Total', 'Transporter', 'Status'])
    style_header(ws)

    for po in purchases:
        first_item = True
        for item in po.items:
            ws.append([
                po.order_number or str(po.id) if first_item else '',
                str(po.order_date) if first_item else '',
                po.supplier.name if first_item else '',
                item.variant.product.name,
                item.variant.display_name,
                item.variant.sku_code,
                item.quantity_dispatched,
                item.quantity_received,
                item.unit_price,
                item.gst_percent,
                item.gst_amount,
                item.total_amount,
                po.transporter or '' if first_item else '',
                po.status if first_item else '',
            ])
            row = ws.max_row
            for col in [9, 11, 12]:
                ws.cell(row=row, column=col).number_format = CURRENCY_FMT
            first_item = False

    ws.append([])
    summary_row = ws.max_row + 1
    ws.append(['', '', '', '', '', 'TOTAL', '', '', '', '', '',
               sum(po.total_amount for po in purchases)])
    ws.cell(row=summary_row, column=12).number_format = CURRENCY_FMT
    ws.cell(row=summary_row, column=12).font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=6).font = Font(bold=True, size=11)

    auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='Sportspedal_Purchases.xlsx', as_attachment=True)


def export_inventory(inventory_data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    ws.append(['Product', 'Category', 'Color', 'Size', 'SKU',
               'Inward', 'Outward', 'Stock', 'Cost Price', 'Stock Value'])
    style_header(ws)

    total_stock = 0
    total_value = 0
    for item in inventory_data:
        ws.append([
            item['product_name'],
            item['category'],
            item['color'],
            item['size'],
            item['sku'],
            item['inward'],
            item['outward'],
            item['stock'],
            item['cost_price'],
            item['stock_value'],
        ])
        row = ws.max_row
        ws.cell(row=row, column=9).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10).number_format = CURRENCY_FMT
        # Color coding
        if item['stock'] <= 0:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='f8d7da')
        elif item['stock'] < 3:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='fff3cd')
        total_stock += item['stock']
        total_value += item['stock_value']

    ws.append([])
    summary_row = ws.max_row + 1
    ws.append(['', '', '', '', 'TOTAL', '', '', total_stock, '', total_value])
    ws.cell(row=summary_row, column=5).font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=8).font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=10).font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=10).number_format = CURRENCY_FMT

    auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='Sportspedal_Inventory.xlsx', as_attachment=True)
